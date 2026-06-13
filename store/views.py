import json
import hmac
import hashlib
import base64
import urllib.request
import urllib.error
from functools import wraps
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseForbidden
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.models import User
from django.conf import settings
from django.core.cache import cache
from django.db.models import F
from django.db import transaction
from django.contrib.auth.models import User as AuthUser
from django.db.models import Q
from .models import Order, Product, OfferGroup, Review, SiteVideo, ProductImage, Category, Cart, CartItem, Coupon, BlogPost, OrderStatusLog


# ── Rate limiting ────────────────────────────────────────────────────────────

def _is_rate_limited(key, max_hits, window_seconds):
    """Returns True when the caller has exceeded max_hits in window_seconds."""
    count = cache.get(key, 0)
    if count >= max_hits:
        return True
    cache.set(key, count + 1, window_seconds)
    return False


def _client_ip(request):
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    return forwarded.split(',')[0].strip() if forwarded else request.META.get('REMOTE_ADDR', '')


# ── Upload validation ─────────────────────────────────────────────────────────

ALLOWED_IMAGE_EXT  = {'.jpg', '.jpeg', '.png', '.webp'}
ALLOWED_IMAGE_MIME = {'image/jpeg', 'image/png', 'image/webp', 'image/jpg'}
MAX_IMAGE_MB       = 5

def _parse_grams(label):
    """Extract a gram quantity from a weight label like '500g', '250 g', '1kg'."""
    import re
    if not label:
        return 0
    m = re.search(r'([\d.]+)\s*(kg|g)?', str(label).lower())
    if not m:
        return 0
    try:
        val = float(m.group(1))
    except ValueError:
        return 0
    return val * 1000 if m.group(2) == 'kg' else val


def _validate_image(f, max_mb=MAX_IMAGE_MB):
    """Return an error string if `f` isn't an acceptable image, else None.
    Checks extension whitelist + size cap + that it actually decodes as an image
    (defeats a non-image file renamed to .jpg)."""
    if not f:
        return None
    import os
    ext = os.path.splitext(f.name)[1].lower()
    if ext not in ALLOWED_IMAGE_EXT:
        return 'Only JPG, PNG or WEBP images are allowed.'
    if f.size > max_mb * 1024 * 1024:
        return f'Image must be under {max_mb} MB.'
    ctype = getattr(f, 'content_type', '') or ''
    if ctype and ctype.lower() not in ALLOWED_IMAGE_MIME:
        return 'Invalid image type.'
    try:
        from PIL import Image
        f.seek(0)
        Image.open(f).verify()   # raises if not a valid image
        f.seek(0)
    except Exception:
        return 'That file is not a valid image.'
    return None


# ── Razorpay helpers ──────────────────────────────────────────────────────────

def _rp_create_order(amount_paise, receipt):
    """Create a Razorpay order via REST API (no external package needed)."""
    creds = base64.b64encode(
        f'{settings.RAZORPAY_KEY_ID}:{settings.RAZORPAY_KEY_SECRET}'.encode()
    ).decode()
    payload = json.dumps({'amount': amount_paise, 'currency': 'INR', 'receipt': receipt}).encode()
    req = urllib.request.Request(
        'https://api.razorpay.com/v1/orders',
        data=payload,
        headers={'Authorization': f'Basic {creds}', 'Content-Type': 'application/json'},
    )
    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read())


def _rp_refund(payment_id, amount_paise):
    """Issue a Razorpay refund via REST API. Returns (refund_dict, error_string)."""
    creds = base64.b64encode(
        f'{settings.RAZORPAY_KEY_ID}:{settings.RAZORPAY_KEY_SECRET}'.encode()
    ).decode()
    payload = json.dumps({'amount': amount_paise}).encode()
    req = urllib.request.Request(
        f'https://api.razorpay.com/v1/payments/{payment_id}/refund',
        data=payload,
        headers={'Authorization': f'Basic {creds}', 'Content-Type': 'application/json'},
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            return json.loads(resp.read()), None
    except urllib.error.HTTPError as e:
        try:
            body = json.loads(e.read())
            return None, body.get('error', {}).get('description', str(e))
        except Exception:
            return None, f'HTTP {e.code}'
    except Exception as e:
        return None, str(e)


def _log_status(order, to_status, changed_by=None, note='', from_status=None):
    """Record a status transition. from_status defaults to the order's current status."""
    try:
        OrderStatusLog.objects.create(
            order=order,
            from_status=from_status if from_status is not None else (order.status or ''),
            to_status=to_status,
            changed_by=changed_by if (changed_by and changed_by.is_authenticated) else None,
            note=note,
        )
    except Exception:
        pass  # logging must never break the actual transition


def _rp_verify(rp_order_id, rp_payment_id, rp_signature):
    """Verify Razorpay payment signature using HMAC-SHA256."""
    secret = settings.RAZORPAY_KEY_SECRET
    if not secret:
        return False
    msg = f'{rp_order_id}|{rp_payment_id}'.encode()
    expected = hmac.new(secret.encode(), msg, hashlib.sha256).hexdigest()
    return hmac.compare_digest(expected, rp_signature)


def _rp_verify_webhook(raw_body, signature):
    """Verify Razorpay webhook signature (uses webhook secret, not API key secret)."""
    secret = settings.RAZORPAY_WEBHOOK_SECRET
    if not secret:
        return False
    expected = hmac.new(secret.encode(), raw_body, hashlib.sha256).hexdigest()
    return hmac.compare_digest(expected, signature)


def _confirm_order_paid(razorpay_order_id, razorpay_payment_id, amount_paise):
    """Mark an order paid. Called by both verify_payment and the webhook — idempotent."""
    order = Order.objects.filter(razorpay_order_id=razorpay_order_id).first()
    if not order:
        return
    # Verify the amount matches to prevent underpayment fraud
    expected_paise = int(float(order.total) * 100)
    if amount_paise != expected_paise:
        return
    if order.payment_status == 'paid':
        return  # already confirmed — do nothing
    prev_status               = order.status
    order.payment_status      = 'paid'
    order.payment_method      = 'online'
    order.razorpay_payment_id = razorpay_payment_id
    order.status              = 'confirmed'
    order.save(update_fields=['payment_status', 'payment_method', 'razorpay_payment_id', 'status'])
    _log_status(order, 'confirmed', None, note='Payment captured', from_status=prev_status)


# ── helpers ──────────────────────────────────────────────────────────────────

ADMIN_OTP_TTL          = 8 * 3600          # re-verify admin email-OTP every 8 hours
ADMIN_OTP_SESSION_KEY  = 'admin_otp_ok_until'


def _admin_otp_ok(request):
    import time
    return request.session.get(ADMIN_OTP_SESSION_KEY, 0) > time.time()


def _staff_required(func):
    """Gate: must be authenticated + staff + email-OTP verified for this session."""
    @wraps(func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect(f'/admin/login/?next={request.path}')
        if not request.user.is_staff:
            return HttpResponseForbidden(
                '<h2 style="font-family:sans-serif;padding:40px">403 — Staff access only. '
                '<a href="/admin/login/">Login as admin</a></h2>'
            )
        if not _admin_otp_ok(request):
            # AJAX/JSON callers get 403 (the admin page will already be verified
            # before any AJAX runs); page loads get redirected to the OTP gate.
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': 'Admin verification required', 'verify': '/manage/verify/'}, status=403)
            return redirect(f'/manage/verify/?next={request.path}')
        return func(request, *args, **kwargs)
    return wrapper


def _pending_orders():
    return Order.objects.filter(status='pending').count()


def _send_admin_otp(user, otp):
    """Email a staff member their admin-access OTP. Sent off-thread so the verify
    page never blocks on the email call (use `manage.py admin_otp` to break-glass
    if email delivery is down)."""
    import threading

    def _send():
        try:
            from django.core.mail import send_mail
            send_mail(
                subject='Your Malola admin verification code',
                message=(f'Your admin access code is: {otp}\n\n'
                         f'It is valid for 10 minutes. If you did not try to access the '
                         f'Malola admin panel, change your password immediately.'),
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
                recipient_list=[user.email],
                fail_silently=True,
            )
        except Exception:
            pass

    threading.Thread(target=_send, daemon=True).start()


def manage_verify(request):
    """Email-OTP gate for the /manage admin area. Not wrapped in _staff_required
    (it must be reachable to clear the gate) but enforces staff itself."""
    import time, random
    if not request.user.is_authenticated:
        return redirect('/admin/login/?next=/manage/')
    if not request.user.is_staff:
        return HttpResponseForbidden(
            '<h2 style="font-family:sans-serif;padding:40px">403 — Staff access only.</h2>')

    nxt = request.POST.get('next') or request.GET.get('next') or '/manage/'
    if not nxt.startswith('/manage'):   # prevent open-redirect
        nxt = '/manage/'
    if _admin_otp_ok(request):
        return redirect(nxt)

    cache_key    = f'admin_otp:{request.user.id}'
    attempts_key = f'admin_otp_attempts:{request.user.id}'
    error = None
    sent  = False

    if request.method == 'POST':
        if request.POST.get('action') == 'send':
            if _is_rate_limited(f'admin_otp_send:{request.user.id}', 5, 600):
                error = 'Too many code requests. Please wait a few minutes.'
            else:
                otp = str(random.randint(100000, 999999))
                cache.set(cache_key, otp, timeout=600)
                cache.delete(attempts_key)
                _send_admin_otp(request.user, otp)
                sent = True
        else:
            otp_in = request.POST.get('otp', '').strip()
            stored = cache.get(cache_key)
            if _is_rate_limited(f'admin_otp_verify:{request.user.id}', 10, 600):
                error = 'Too many attempts. Please wait a few minutes.'
            elif not stored:
                error = 'Code expired. Request a new one.'
            elif stored != otp_in:
                attempts = cache.get(attempts_key, 0) + 1
                cache.set(attempts_key, attempts, timeout=600)
                if attempts >= 5:
                    cache.delete(cache_key); cache.delete(attempts_key)
                    error = 'Too many incorrect codes. Request a new one.'
                else:
                    error = 'Incorrect code. Please try again.'
            else:
                cache.delete(cache_key); cache.delete(attempts_key)
                request.session[ADMIN_OTP_SESSION_KEY] = int(time.time()) + ADMIN_OTP_TTL
                return redirect(nxt)
    else:
        # Auto-send a code on first arrival (if none active and not rate-limited)
        if not cache.get(cache_key) and not _is_rate_limited(f'admin_otp_send:{request.user.id}', 5, 600):
            otp = str(random.randint(100000, 999999))
            cache.set(cache_key, otp, timeout=600)
            _send_admin_otp(request.user, otp)
            sent = True

    # Mask the email for display: a***@gmail.com
    em = request.user.email or ''
    masked = em
    if '@' in em:
        local, dom = em.split('@', 1)
        masked = (local[0] + '***') if local else '***'
        masked += '@' + dom
    return render(request, 'store/admin_verify.html', {
        'error': error, 'sent': sent, 'next': nxt, 'email_masked': masked,
    })


# ── public pages ─────────────────────────────────────────────────────────────

_HOME_TTL = 120  # 2 minutes

def _bust_product_cache():
    """Call after any admin create/edit/delete so the front-end cache refreshes immediately."""
    for key in ('home_our_products', 'home_new_arrivals', 'home_whats_new',
                'home_reviews', 'home_videos', 'home_categories',
                'shop_our_products', 'shop_new_arrivals', 'shop_categories', 'shop_sales_map'):
        cache.delete(key)

def home(request):
    db_our_products = cache.get('home_our_products')
    if db_our_products is None:
        db_our_products = list(Product.objects.filter(category='our_products', is_active=True))
        cache.set('home_our_products', db_our_products, _HOME_TTL)

    db_new_arrivals = cache.get('home_new_arrivals')
    if db_new_arrivals is None:
        db_new_arrivals = list(Product.objects.filter(category='new_arrival', is_active=True))
        cache.set('home_new_arrivals', db_new_arrivals, _HOME_TTL)

    db_whats_new = cache.get('home_whats_new')
    if db_whats_new is None:
        db_whats_new = list(Product.objects.filter(is_active=True).order_by('-created_at')[:4])
        cache.set('home_whats_new', db_whats_new, _HOME_TTL)

    db_reviews = cache.get('home_reviews')
    if db_reviews is None:
        db_reviews = list(Review.objects.filter(is_active=True))
        cache.set('home_reviews', db_reviews, _HOME_TTL)

    db_videos = cache.get('home_videos')
    if db_videos is None:
        db_videos = list(SiteVideo.objects.filter(is_active=True))
        cache.set('home_videos', db_videos, _HOME_TTL)

    db_categories = cache.get('home_categories')
    if db_categories is None:
        db_categories = list(Category.objects.filter(is_active=True).order_by('sort_order', 'name'))
        cache.set('home_categories', db_categories, _HOME_TTL)

    return render(request, 'store/index.html', {
        'db_our_products': db_our_products,
        'db_new_arrivals': db_new_arrivals,
        'db_whats_new':    db_whats_new,
        'db_reviews':      db_reviews,
        'db_videos':       db_videos,
        'db_categories':   db_categories,
    })


SHOP_FILTERS = [
    {'group': 'Stage',    'options': ['Family', 'Kids']},
    {'group': 'Diet',     'options': ['Dairy Free', 'Gluten Free Grains', 'Nut Free', 'Salt Free', 'Sugar Free']},
    {'group': 'Grain',    'options': ['Foxtail Millet', 'Jowar', 'Little Millet', 'Oat', 'Ragi']},
    {'group': 'Benefits', 'options': ['Fibre Rich', 'Healthy Weight Gain', 'Protein Rich', 'Travel Friendly']},
]


def shop(request):
    from django.db.models import Count
    from collections import defaultdict

    db_our_products = cache.get('shop_our_products')
    if db_our_products is None:
        db_our_products = list(Product.objects.filter(category='our_products', is_active=True))
        cache.set('shop_our_products', db_our_products, _HOME_TTL)

    db_new_arrivals = cache.get('shop_new_arrivals')
    if db_new_arrivals is None:
        db_new_arrivals = list(Product.objects.filter(category='new_arrival', is_active=True))
        cache.set('shop_new_arrivals', db_new_arrivals, _HOME_TTL)

    # Units sold per product slug (cached — avoids full Order scan every request)
    sales = cache.get('shop_sales_map')
    if sales is None:
        sales = defaultdict(int)
        for o in Order.objects.filter(payment_status__in=['paid', 'cod']).only('items'):
            for it in (o.items or []):
                try:
                    sales[it.get('slug', '')] += int(it.get('qty', 1))
                except (TypeError, ValueError):
                    pass
        cache.set('shop_sales_map', dict(sales), 300)  # cache 5 minutes
    for p in db_our_products + db_new_arrivals:
        p.units_sold = sales.get(p.slug, 0)

    categories = cache.get('shop_categories')
    if categories is None:
        categories = list(
            Category.objects
            .filter(is_active=True)
            .annotate(product_count=Count('products'))
            .filter(product_count__gt=0)
            .order_by('sort_order', 'name')
        )
        cache.set('shop_categories', categories, _HOME_TTL)

    db_whats_new = cache.get('home_whats_new') or list(
        Product.objects.filter(is_active=True).order_by('-created_at')[:4]
    )
    return render(request, 'store/shop.html', {
        'db_our_products': db_our_products,
        'db_new_arrivals': db_new_arrivals,
        'categories':      categories,
        'shop_filters':    SHOP_FILTERS,
        'db_whats_new':    db_whats_new,
    })


def product(request):
    db_new_arrivals = Product.objects.filter(category='new_arrival', is_active=True)
    db_our_products = Product.objects.filter(category='our_products', is_active=True)
    pid        = request.GET.get('id', '')
    db_product = None
    db_ingredients_json = '[]'
    db_weights_json     = '["100g","200g","500g"]'
    db_nutrition_json   = '{}'
    db_gallery_json     = '[]'
    db_product_json = 'null'
    if pid:
        db_product = Product.objects.filter(slug=pid, is_active=True).first()
        if db_product:
            ingredients = [i.strip() for i in db_product.ingredients.split(',') if i.strip()] if db_product.ingredients else []
            weights     = [w.strip() for w in db_product.weights.split(',') if w.strip()] if db_product.weights else ['100g', '200g', '500g']
            nutrition   = {}
            if db_product.nut_calories:      nutrition['Calories']      = db_product.nut_calories
            if db_product.nut_protein:       nutrition['Protein']       = db_product.nut_protein
            if db_product.nut_fat:           nutrition['Fat']           = db_product.nut_fat
            if db_product.nut_saturated_fat: nutrition['Saturated Fat'] = db_product.nut_saturated_fat
            if db_product.nut_trans_fat:     nutrition['Trans Fat']     = db_product.nut_trans_fat
            if db_product.nut_carbs:         nutrition['Carbs']         = db_product.nut_carbs
            if db_product.nut_fibre:         nutrition['Fibre']         = db_product.nut_fibre
            if db_product.nut_sugar:         nutrition['Sugar']         = db_product.nut_sugar
            if db_product.nut_sodium:        nutrition['Sodium']        = db_product.nut_sodium
            gallery = [img.image.url for img in db_product.gallery_images.all()]
            recipe = None
            if db_product.recipe_name:
                recipe = {
                    'name':         db_product.recipe_name,
                    'prepTime':     db_product.recipe_prep_time,
                    'cookTime':     db_product.recipe_cook_time,
                    'servings':     db_product.recipe_servings,
                    'ingredients':  db_product.recipe_ingredients,
                    'instructions': db_product.recipe_instructions,
                    'image':        db_product.recipe_image.url if db_product.recipe_image else '',
                    'videoUrl':     db_product.recipe_video_url,
                }
            data = {
                'id':            db_product.slug,
                'name':          db_product.title,
                'category':      db_product.product_type or db_product.get_category_display(),
                'categoryId':    db_product.category,
                'price':         float(db_product.price),
                'discountPrice': float(db_product.discount_price) if db_product.discount_price else None,
                'badge':         db_product.badge,
                'badgeColor':    db_product.badge_color,
                'bg':            db_product.card_bg,
                'image':         db_product.image.url if db_product.image else '',
                'imageBack':     db_product.image.url if db_product.image else '',
                'description':   db_product.description,
                'ingredients':   ingredients,
                'nutrition':     nutrition,
                'weights':       weights,
                'galleryImages': gallery,
                'rating':        float(db_product.rating),
                'reviewCount':   db_product.reviews_count,
                'stockQuantity': db_product.stock_quantity,
                'brand':         db_product.brand,
                'sku':           db_product.sku,
                'shortDesc':     db_product.short_description,
                'healthBenefits':db_product.health_benefits,
                'certifications':{
                    'organic': db_product.cert_organic,
                    'nonGmo':  db_product.cert_non_gmo,
                    'vegan':   db_product.cert_vegan,
                    'halal':   db_product.cert_halal,
                    'iso':     db_product.cert_iso,
                },
                'recipe': recipe,
                'productInfo': {
                    'packageSize':         db_product.package_size,
                    'countryOfOrigin':     db_product.country_of_origin,
                    'shelfLife':           db_product.shelf_life,
                    'storageInstructions': db_product.storage_instructions,
                    'manufacturerDetails': db_product.manufacturer_details,
                },
            }
            db_product_json = json.dumps(data)
    db_whats_new = Product.objects.filter(is_active=True).order_by('-created_at')[:4]
    return render(request, 'store/product.html', {
        'db_new_arrivals':  db_new_arrivals,
        'db_our_products':  db_our_products,
        'db_product':       db_product,
        'db_product_json':  db_product_json,
        'db_whats_new':     db_whats_new,
    })


def orders_page(request):
    if not request.user.is_authenticated:
        return redirect('/?login=1')
    orders = Order.objects.filter(user=request.user)
    ctx = _nav_ctx()
    ctx['orders'] = orders
    return render(request, 'store/orders.html', ctx)


# ── admin product manager ─────────────────────────────────────────────────────

@_staff_required
def manage_products(request):
    all_products = Product.objects.all()
    stats = {
        'total':  all_products.count(),
        'our':    all_products.filter(category='our_products').count(),
        'new':    all_products.filter(category='new_arrival').count(),
        'active': all_products.filter(is_active=True).count(),
    }
    paginator = Paginator(all_products, 20)
    products  = paginator.get_page(request.GET.get('page'))
    return render(request, 'store/admin_panel.html', {
        'products':       products,
        'stats':          stats,
        'pending_orders': _pending_orders(),
    })


def _product_from_post(post, files, product=None):
    """Pull all product fields from POST/FILES. Returns (data_dict, error_string)."""
    title        = post.get('title', '').strip()
    slug         = post.get('slug', '').strip()
    description  = post.get('description', '').strip()
    price        = post.get('price', '').strip()
    category     = post.get('category', 'our_products')
    badge        = post.get('badge', '').strip()
    badge_color  = post.get('badge_color', '#1565C0')
    card_bg      = post.get('card_bg', '#f4f7ff')
    is_active    = post.get('is_active') == 'on'
    image        = files.get('image')
    video        = files.get('video')
    product_type = post.get('product_type', '').strip()
    weights      = post.get('weights', '100g,200g,500g').strip()
    ingredients  = post.get('ingredients', '').strip()
    nut_calories = post.get('nut_calories', '').strip()
    nut_protein  = post.get('nut_protein', '').strip()
    nut_fat      = post.get('nut_fat', '').strip()
    nut_carbs    = post.get('nut_carbs', '').strip()
    nut_fibre    = post.get('nut_fibre', '').strip()
    # extended fields
    short_description    = post.get('short_description', '').strip()
    brand                = post.get('brand', '').strip()
    sku                  = post.get('sku', '').strip()
    discount_price_raw   = post.get('discount_price', '').strip()
    discount_price       = None
    if discount_price_raw:
        try:
            discount_price = float(discount_price_raw)
        except ValueError:
            pass
    try:
        stock_quantity = int(post.get('stock_quantity', '0') or '0')
    except ValueError:
        stock_quantity = 0
    track_inventory = post.get('track_inventory') == 'on'
    try:
        max_order_qty = max(0, int(post.get('max_order_qty', '0') or '0'))
    except ValueError:
        max_order_qty = 0
    try:
        gst_rate = max(0.0, float(post.get('gst_rate', '5') or '5'))
    except ValueError:
        gst_rate = 5.0
    package_size         = post.get('package_size', '').strip()
    country_of_origin    = post.get('country_of_origin', '').strip()
    shelf_life           = post.get('shelf_life', '').strip()
    storage_instructions = post.get('storage_instructions', '').strip()
    manufacturer_details = post.get('manufacturer_details', '').strip()
    nut_saturated_fat    = post.get('nut_saturated_fat', '').strip()
    nut_trans_fat        = post.get('nut_trans_fat', '').strip()
    nut_sugar            = post.get('nut_sugar', '').strip()
    nut_sodium           = post.get('nut_sodium', '').strip()
    recipe_name          = post.get('recipe_name', '').strip()
    recipe_prep_time     = post.get('recipe_prep_time', '').strip()
    recipe_cook_time     = post.get('recipe_cook_time', '').strip()
    recipe_servings      = post.get('recipe_servings', '').strip()
    recipe_ingredients   = post.get('recipe_ingredients', '').strip()
    recipe_instructions  = post.get('recipe_instructions', '').strip()
    recipe_image         = files.get('recipe_image')
    recipe_video_url     = post.get('recipe_video_url', '').strip()
    health_benefits      = post.get('health_benefits', '').strip()
    cert_organic         = post.get('cert_organic') == 'on'
    cert_non_gmo         = post.get('cert_non_gmo') == 'on'
    cert_vegan           = post.get('cert_vegan') == 'on'
    cert_halal           = post.get('cert_halal') == 'on'
    cert_iso             = post.get('cert_iso') == 'on'
    seo_title            = post.get('seo_title', '').strip()
    seo_description      = post.get('seo_description', '').strip()
    seo_keywords         = post.get('seo_keywords', '').strip()
    tags                 = post.get('tags', '').strip()
    try:
        rating = float(post.get('rating', '4.5') or '4.5')
        rating = max(0.0, min(5.0, rating))
    except ValueError:
        rating = 4.5
    try:
        reviews_count = int(post.get('reviews_count', '0') or '0')
    except ValueError:
        reviews_count = 0
    offer_group_id = post.get('offer_group', '').strip()
    offer_group = None
    if offer_group_id:
        try:
            offer_group = OfferGroup.objects.get(pk=int(offer_group_id))
        except (OfferGroup.DoesNotExist, ValueError):
            pass

    product_category_id = post.get('product_category', '').strip()
    product_category = None
    if product_category_id:
        try:
            product_category = Category.objects.get(pk=int(product_category_id))
        except (Category.DoesNotExist, ValueError):
            pass

    if not all([title, slug, description, price]):
        return None, 'Title, slug, description and price are all required.'
    if not image and not product:
        return None, 'A product image is required.'
    for _f in (image, recipe_image):
        _err = _validate_image(_f)
        if _err:
            return None, _err

    data = dict(
        title=title, slug=slug, description=description, price=price,
        category=category, product_category=product_category,
        badge=badge, badge_color=badge_color, card_bg=card_bg,
        is_active=is_active, product_type=product_type, weights=weights,
        ingredients=ingredients, nut_calories=nut_calories, nut_protein=nut_protein,
        nut_fat=nut_fat, nut_carbs=nut_carbs, nut_fibre=nut_fibre,
        rating=rating, reviews_count=reviews_count, offer_group=offer_group,
        short_description=short_description, brand=brand, sku=sku,
        discount_price=discount_price, stock_quantity=stock_quantity,
        track_inventory=track_inventory, max_order_qty=max_order_qty, gst_rate=gst_rate,
        package_size=package_size, country_of_origin=country_of_origin,
        shelf_life=shelf_life, storage_instructions=storage_instructions,
        manufacturer_details=manufacturer_details,
        nut_saturated_fat=nut_saturated_fat, nut_trans_fat=nut_trans_fat,
        nut_sugar=nut_sugar, nut_sodium=nut_sodium,
        recipe_name=recipe_name, recipe_prep_time=recipe_prep_time,
        recipe_cook_time=recipe_cook_time, recipe_servings=recipe_servings,
        recipe_ingredients=recipe_ingredients, recipe_instructions=recipe_instructions,
        recipe_video_url=recipe_video_url,
        health_benefits=health_benefits,
        cert_organic=cert_organic, cert_non_gmo=cert_non_gmo,
        cert_vegan=cert_vegan, cert_halal=cert_halal, cert_iso=cert_iso,
        seo_title=seo_title, seo_description=seo_description, seo_keywords=seo_keywords,
        tags=tags,
    )
    if image:
        data['image'] = image
    if video:
        data['video'] = video
    if recipe_image:
        data['recipe_image'] = recipe_image
    return data, None


@_staff_required
def manage_add_product(request):
    error = None
    if request.method == 'POST':
        data, error = _product_from_post(request.POST, request.FILES)
        if not error:
            slug = data['slug']
            if Product.objects.filter(slug=slug).exists():
                error = f'A product with the slug "{slug}" already exists. Choose a different one.'
            else:
                gallery_files = request.FILES.getlist('gallery_images')
                error = next((_validate_image(gf) for gf in gallery_files if _validate_image(gf)), None)
                if not error:
                    try:
                        product = Product.objects.create(**data)
                        for gf in gallery_files:
                            ProductImage.objects.create(product=product, image=gf)
                        _bust_product_cache()
                        return redirect('manage_products')
                    except Exception as exc:
                        error = str(exc)
    return render(request, 'store/admin_form.html', {
        'action': 'Add', 'error': error,
        'offer_groups': OfferGroup.objects.filter(is_active=True),
        'categories': Category.objects.filter(is_active=True),
        'pending_orders': _pending_orders(),
    })


@_staff_required
def manage_edit_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    error = None
    if request.method == 'POST':
        data, error = _product_from_post(request.POST, request.FILES, product)
        gallery_files = request.FILES.getlist('gallery_images')
        if not error:
            error = next((_validate_image(gf) for gf in gallery_files if _validate_image(gf)), None)
        if not error:
            for field, value in data.items():
                setattr(product, field, value)
            try:
                product.save()
                for gf in gallery_files:
                    ProductImage.objects.create(product=product, image=gf)
                _bust_product_cache()
                return redirect('manage_products')
            except Exception as exc:
                error = str(exc)

    return render(request, 'store/admin_form.html', {
        'action': 'Edit',
        'product': product,
        'error': error,
        'offer_groups': OfferGroup.objects.filter(is_active=True),
        'categories': Category.objects.filter(is_active=True),
        'gallery_images': product.gallery_images.all(),
        'pending_orders': _pending_orders(),
    })


@_staff_required
@require_POST
def manage_delete_gallery_image(request, pk, img_pk):
    product = get_object_or_404(Product, pk=pk)
    img = get_object_or_404(ProductImage, pk=img_pk, product=product)
    if img.image:
        img.image.delete(save=False)
    img.delete()
    return redirect(f'/manage/edit/{pk}/#media')


@_staff_required
def manage_delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        _bust_product_cache()
    return redirect('manage_products')


# ── admin categories ─────────────────────────────────────────────────────────

@_staff_required
def manage_categories(request):
    from django.db.models import Count
    categories = Category.objects.annotate(
        product_count=Count('products')
    ).order_by('sort_order', 'name')
    return render(request, 'store/admin_categories.html', {
        'categories': categories,
        'pending_orders': _pending_orders(),
    })


@_staff_required
def manage_add_category(request):
    error = None
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        sort_order = request.POST.get('sort_order', '0').strip()
        is_active = request.POST.get('is_active') == 'on'
        if not name:
            error = 'Category name is required.'
        elif Category.objects.filter(name=name).exists():
            error = f'A category named "{name}" already exists.'
        else:
            try:
                cat = Category(
                    name=name, description=description,
                    sort_order=int(sort_order or 0), is_active=is_active,
                )
                if 'image' in request.FILES:
                    cat.image = request.FILES['image']
                cat.save()
                _bust_product_cache()
                return redirect('manage_categories')
            except Exception as exc:
                error = str(exc)
    return render(request, 'store/admin_category_form.html', {
        'action': 'Add', 'error': error, 'pending_orders': _pending_orders(),
    })


@_staff_required
def manage_edit_category(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    error = None
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        sort_order = request.POST.get('sort_order', '0').strip()
        is_active = request.POST.get('is_active') == 'on'
        if not name:
            error = 'Category name is required.'
        elif Category.objects.filter(name=name).exclude(pk=pk).exists():
            error = f'A category named "{name}" already exists.'
        else:
            cat.name = name
            cat.description = description
            cat.sort_order = int(sort_order or 0)
            cat.is_active = is_active
            if 'image' in request.FILES:
                if cat.image:
                    cat.image.delete(save=False)
                cat.image = request.FILES['image']
            elif request.POST.get('image_clear') and cat.image:
                cat.image.delete(save=False)
                cat.image = None
            cat.save()
            _bust_product_cache()
            return redirect('manage_categories')
    return render(request, 'store/admin_category_form.html', {
        'action': 'Edit', 'cat': cat, 'error': error, 'pending_orders': _pending_orders(),
    })


@_staff_required
def manage_delete_category(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        cat.delete()
    return redirect('manage_categories')


# ── admin orders ──────────────────────────────────────────────────────────────

def _orders_queryset(request):
    """Return filtered Order queryset based on GET params — shared by list + exports."""
    qs = Order.objects.select_related('user').order_by('-created_at')
    status_filter  = request.GET.get('status', '').strip()
    payment_filter = request.GET.get('payment', '').strip()
    search         = request.GET.get('q', '').strip()
    date_from      = request.GET.get('date_from', '').strip()
    date_to        = request.GET.get('date_to', '').strip()
    if status_filter:
        qs = qs.filter(status=status_filter)
    if payment_filter:
        qs = qs.filter(payment_status=payment_filter)
    if search:
        q_obj = Q(name__icontains=search) | Q(phone__icontains=search)
        if search.isdigit():
            q_obj |= Q(id=int(search))
        qs = qs.filter(q_obj)
    if date_from:
        try:
            qs = qs.filter(created_at__date__gte=date_from)
        except Exception:
            pass
    if date_to:
        try:
            qs = qs.filter(created_at__date__lte=date_to)
        except Exception:
            pass
    return qs


@_staff_required
def manage_orders(request):
    from django.db.models import Sum
    from urllib.parse import urlencode
    qs = _orders_queryset(request)
    total_revenue = Order.objects.filter(
        payment_status__in=['paid', 'cod']
    ).aggregate(rev=Sum('total'))['rev'] or 0
    stats = {
        'total':   Order.objects.count(),
        'pending': Order.objects.filter(status='pending').count(),
        'paid':    Order.objects.filter(payment_status='paid').count(),
        'cod':     Order.objects.filter(payment_status='cod').count(),
        'revenue': round(total_revenue, 2),
    }
    paginator = Paginator(qs, 25)
    orders    = paginator.get_page(request.GET.get('page'))
    status_filter  = request.GET.get('status', '')
    payment_filter = request.GET.get('payment', '')
    search         = request.GET.get('q', '')
    date_from      = request.GET.get('date_from', '')
    date_to        = request.GET.get('date_to', '')
    filter_params  = {k: v for k, v in {
        'q': search, 'status': status_filter, 'payment': payment_filter,
        'date_from': date_from, 'date_to': date_to,
    }.items() if v}
    filter_qs = urlencode(filter_params)
    return render(request, 'store/admin_orders.html', {
        'orders':         orders,
        'status_filter':  status_filter,
        'payment_filter': payment_filter,
        'search':         search,
        'date_from':      date_from,
        'date_to':        date_to,
        'filter_qs':      filter_qs,
        'status_choices': Order.STATUS_CHOICES,
        'payment_choices': Order.PAYMENT_STATUS,
        'stats':          stats,
        'pending_orders': stats['pending'],
    })


@_staff_required
def export_orders_csv(request):
    import csv
    from django.http import HttpResponse
    qs = _orders_queryset(request)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="malola_orders.csv"'
    writer = csv.writer(response)
    writer.writerow(['Order ID', 'Date', 'Customer', 'Phone', 'Address',
                     'Items', 'Coupon', 'Discount', 'Total', 'Payment', 'Status'])
    for o in qs:
        items_str = '; '.join(
            f"{i.get('name','')} x{i.get('qty',1)} @₹{i.get('price','')}"
            for i in (o.items or [])
        )
        writer.writerow([
            o.id, o.created_at.strftime('%d/%m/%Y %H:%M'),
            o.name, o.phone, o.address, items_str,
            o.coupon_code or '', o.discount_amount or 0,
            o.total, o.get_payment_status_display(), o.get_status_display(),
        ])
    return response


@_staff_required
def export_orders_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse('openpyxl not installed. Run: pip install openpyxl', status=500)
    from django.http import HttpResponse
    qs = _orders_queryset(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Orders'
    header_fill = PatternFill('solid', fgColor='0D2B6B')
    header_font = Font(bold=True, color='FECF0A', size=10)
    headers = ['Order ID', 'Date', 'Customer', 'Phone', 'Address',
               'Items', 'Coupon', 'Discount (₹)', 'Total (₹)', 'Payment', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for o in qs:
        items_str = '; '.join(
            f"{i.get('name','')} x{i.get('qty',1)}"
            for i in (o.items or [])
        )
        ws.append([
            o.id, o.created_at.strftime('%d/%m/%Y %H:%M'),
            o.name, o.phone, o.address, items_str,
            o.coupon_code or '', float(o.discount_amount or 0),
            float(o.total), o.get_payment_status_display(), o.get_status_display(),
        ])
    col_widths = [10, 18, 20, 14, 35, 50, 12, 14, 12, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="malola_orders.xlsx"'
    wb.save(response)
    return response


@_staff_required
def export_orders_print(request):
    qs = _orders_queryset(request)
    return render(request, 'store/admin_orders_print.html', {
        'orders': qs,
        'date_from': request.GET.get('date_from', ''),
        'date_to':   request.GET.get('date_to', ''),
        'status_filter':  request.GET.get('status', ''),
        'payment_filter': request.GET.get('payment', ''),
        'search':         request.GET.get('q', ''),
    })


@_staff_required
def manage_order_status(request, pk):
    if request.method == 'POST':
        order = get_object_or_404(Order, pk=pk)
        new_status = request.POST.get('status', '')
        valid = [s[0] for s in Order.STATUS_CHOICES]
        if new_status in valid and new_status != order.status:
            prev = order.status
            order.status = new_status
            order.save(update_fields=['status'])
            _log_status(order, new_status, request.user, note='Status changed by admin', from_status=prev)
    status_back = request.POST.get('status_filter', '')
    if status_back in [s[0] for s in Order.STATUS_CHOICES]:
        return redirect(f'/manage/orders/?status={status_back}')
    return redirect('manage_orders')


@_staff_required
@require_POST
def manage_refund_order(request, pk):
    """Admin partial/full refund: refund selected items×qty via Razorpay, restore stock, update status."""
    order = get_object_or_404(Order, pk=pk)
    if order.payment_status not in ('paid', 'partially_refunded'):
        return JsonResponse({'success': False, 'error': 'This order has no captured online payment to refund.'}, status=400)
    if not order.razorpay_payment_id:
        return JsonResponse({'success': False, 'error': 'No Razorpay payment id on this order.'}, status=400)
    try:
        body    = json.loads(request.body)
        refunds = body.get('items', [])  # [{index, qty}]
        if not refunds:
            return JsonResponse({'success': False, 'error': 'Select at least one item and quantity to refund.'}, status=400)

        with transaction.atomic():
            order = Order.objects.select_for_update().get(pk=pk)
            items = order.items
            # discount ratio so we refund what was actually charged, not the pre-discount MRP
            line_sum = sum(float(i.get('price', 0)) * int(i.get('qty', 1) or 1) for i in items)
            scale    = (float(order.total) / line_sum) if line_sum > 0 else 1.0

            refund_total = 0.0
            restore      = []  # (slug, qty)
            for r in refunds:
                idx = int(r.get('index', -1))
                rq  = int(r.get('qty', 0) or 0)
                if rq <= 0:
                    continue
                if idx < 0 or idx >= len(items):
                    return JsonResponse({'success': False, 'error': 'Invalid item selected.'}, status=400)
                it       = items[idx]
                ordered  = int(it.get('qty', 1) or 1)
                already  = int(it.get('refunded_qty', 0) or 0)
                if already + rq > ordered:
                    return JsonResponse({'success': False, 'error': f"Cannot refund {rq} × \"{it.get('name','')}\" — only {ordered - already} left to refund."}, status=400)
                amt = round(float(it.get('price', 0)) * rq * scale, 2)
                refund_total += amt
                it['refunded_qty']    = already + rq
                it['refunded_amount'] = round(float(it.get('refunded_amount', 0) or 0) + amt, 2)
                restore.append((it.get('slug', ''), rq))

            if refund_total <= 0:
                return JsonResponse({'success': False, 'error': 'Nothing to refund.'}, status=400)

            # Cap so cumulative refund never exceeds what was paid
            remaining = round(float(order.total) - float(order.refunded_amount), 2)
            refund_total = min(refund_total, remaining)

            # Issue the Razorpay refund (still inside the lock; admin volume is low)
            refund, err = _rp_refund(order.razorpay_payment_id, int(round(refund_total * 100)))
            if not (refund and refund.get('id')):
                return JsonResponse({'success': False, 'error': f'Razorpay refund failed: {err}'}, status=502)

            # Persist item changes + restore stock
            order.items = items
            order.refunded_amount = round(float(order.refunded_amount) + refund_total, 2)
            for slug, rq in restore:
                Product.objects.filter(slug=slug, track_inventory=True).update(
                    stock_quantity=F('stock_quantity') + rq
                )
            fully = all(int(i.get('refunded_qty', 0) or 0) >= int(i.get('qty', 1) or 1) for i in items)
            order.payment_status = 'refunded' if fully else 'partially_refunded'
            order.razorpay_refund_id = refund['id']
            order.save(update_fields=['items', 'refunded_amount', 'payment_status', 'razorpay_refund_id'])
            _log_status(order, order.status, request.user,
                        note=f'Refunded ₹{refund_total:.2f} ({order.get_payment_status_display()})',
                        from_status=order.status)

        return JsonResponse({'success': True, 'refunded': refund_total,
                             'payment_status': order.get_payment_status_display(),
                             'total_refunded': float(order.refunded_amount)})
    except Exception as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)


# ── admin offers ──────────────────────────────────────────────────────────────

@_staff_required
def manage_offers(request):
    offers = OfferGroup.objects.prefetch_related('products').all()
    pending_orders = Order.objects.filter(status='pending').count()
    return render(request, 'store/admin_offers.html', {
        'offers': offers,
        'pending_orders': pending_orders,
    })


@_staff_required
def manage_add_offer(request):
    error = None
    if request.method == 'POST':
        name        = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        badge_label = request.POST.get('badge_label', '').strip()
        badge_color = request.POST.get('badge_color', '#c62828')
        is_active   = request.POST.get('is_active') == 'on'
        product_ids = request.POST.getlist('products')
        try:
            discount_pct = max(0, min(100, int(request.POST.get('discount_pct', 0) or 0)))
        except ValueError:
            discount_pct = 0
        if not name:
            error = 'Offer name is required.'
        else:
            offer = OfferGroup.objects.create(
                name=name, description=description, discount_pct=discount_pct,
                badge_label=badge_label, badge_color=badge_color, is_active=is_active,
            )
            if product_ids:
                Product.objects.filter(pk__in=product_ids).update(offer_group=offer)
            return redirect('manage_offers')
    all_products = Product.objects.all().order_by('title')
    return render(request, 'store/admin_offer_form.html', {
        'action': 'Add', 'error': error,
        'all_products': all_products, 'offer': None,
        'pending_orders': _pending_orders(),
    })


@_staff_required
def manage_edit_offer(request, pk):
    offer = get_object_or_404(OfferGroup, pk=pk)
    error = None
    if request.method == 'POST':
        name        = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        badge_label = request.POST.get('badge_label', '').strip()
        badge_color = request.POST.get('badge_color', '#c62828')
        is_active   = request.POST.get('is_active') == 'on'
        product_ids = request.POST.getlist('products')
        try:
            discount_pct = max(0, min(100, int(request.POST.get('discount_pct', 0) or 0)))
        except ValueError:
            discount_pct = 0
        if not name:
            error = 'Offer name is required.'
        else:
            offer.name = name; offer.description = description
            offer.discount_pct = discount_pct; offer.badge_label = badge_label
            offer.badge_color = badge_color; offer.is_active = is_active
            offer.save()
            offer.products.update(offer_group=None)
            if product_ids:
                Product.objects.filter(pk__in=product_ids).update(offer_group=offer)
            return redirect('manage_offers')
    all_products  = Product.objects.all().order_by('title')
    assigned_ids  = list(offer.products.values_list('pk', flat=True))
    return render(request, 'store/admin_offer_form.html', {
        'action': 'Edit', 'error': error,
        'all_products': all_products, 'offer': offer, 'assigned_ids': assigned_ids,
        'pending_orders': _pending_orders(),
    })


@_staff_required
def manage_delete_offer(request, pk):
    offer = get_object_or_404(OfferGroup, pk=pk)
    if request.method == 'POST':
        offer.products.update(offer_group=None)
        offer.delete()
    return redirect('manage_offers')


# ── admin reviews ─────────────────────────────────────────────────────────────

@_staff_required
def manage_reviews(request):
    reviews = Review.objects.select_related('order').all()
    return render(request, 'store/admin_reviews.html', {
        'reviews': reviews,
        'pending_orders': _pending_orders(),
    })


@_staff_required
def manage_delete_review(request, pk):
    review = get_object_or_404(Review, pk=pk)
    if request.method == 'POST':
        review.delete()
    return redirect('manage_reviews')


# ── admin videos ──────────────────────────────────────────────────────────────

@_staff_required
def manage_videos(request):
    videos = SiteVideo.objects.all()
    return render(request, 'store/admin_videos.html', {
        'videos':         videos,
        'pending_orders': _pending_orders(),
    })


@_staff_required
@require_POST
def manage_upload_video(request):
    video_file = request.FILES.get('video')
    title      = request.POST.get('title', '').strip()
    sort_order = request.POST.get('sort_order', '0').strip()
    if not video_file:
        return redirect('manage_videos')
    try:
        sort_order = int(sort_order)
    except ValueError:
        sort_order = 0
    SiteVideo.objects.create(video=video_file, title=title, sort_order=sort_order)
    _bust_product_cache()
    return redirect('manage_videos')


@_staff_required
@require_POST
def manage_delete_video(request, pk):
    video = get_object_or_404(SiteVideo, pk=pk)
    if video.video:
        video.video.delete(save=False)
    video.delete()
    _bust_product_cache()
    return redirect('manage_videos')


@_staff_required
@require_POST
def manage_toggle_video(request, pk):
    video = get_object_or_404(SiteVideo, pk=pk)
    video.is_active = not video.is_active
    video.save(update_fields=['is_active'])
    return redirect('manage_videos')


# ── public blog & about ───────────────────────────────────────────────────────

def blog_list(request):
    posts = BlogPost.objects.filter(is_published=True)
    ctx = _nav_ctx()
    ctx['posts'] = posts
    return render(request, 'store/blog_list.html', ctx)


def blog_detail(request, slug):
    post = get_object_or_404(BlogPost, slug=slug, is_published=True)
    recent = BlogPost.objects.filter(is_published=True).exclude(pk=post.pk)[:4]
    ctx = _nav_ctx()
    ctx.update({'post': post, 'recent': recent})
    return render(request, 'store/blog_detail.html', ctx)


def about_page(request):
    return render(request, 'store/about.html', _nav_ctx())


def _nav_ctx():
    our = cache.get('home_our_products')
    if our is None:
        our = list(Product.objects.filter(category='our_products', is_active=True))
        cache.set('home_our_products', our, _HOME_TTL)
    new = cache.get('home_new_arrivals')
    if new is None:
        new = list(Product.objects.filter(category='new_arrival', is_active=True))
        cache.set('home_new_arrivals', new, _HOME_TTL)
    wn = cache.get('home_whats_new')
    if wn is None:
        wn = list(Product.objects.filter(is_active=True).order_by('-created_at')[:4])
        cache.set('home_whats_new', wn, _HOME_TTL)
    return {'db_our_products': our, 'db_new_arrivals': new, 'db_whats_new': wn}


def shipping_policy(request):
    return render(request, 'store/shipping_policy.html', _nav_ctx())


def privacy_policy(request):
    return render(request, 'store/privacy_policy.html', _nav_ctx())


def terms_page(request):
    return render(request, 'store/terms.html', _nav_ctx())


def refund_policy(request):
    return render(request, 'store/refund_policy.html', _nav_ctx())


def contact_page(request):
    ctx = _nav_ctx()
    if request.method == 'POST':
        ctx['success'] = True
    return render(request, 'store/contact.html', ctx)


def faq_page(request):
    return render(request, 'store/faq.html', _nav_ctx())


def qr_page(request):
    return render(request, 'store/qr.html')


def qr_code_page(request):
    return render(request, 'store/qr_code.html')


# ── admin blog ────────────────────────────────────────────────────────────────

@_staff_required
def manage_blog(request):
    posts = BlogPost.objects.all()
    return render(request, 'store/admin_blog.html', {
        'posts': posts,
        'pending_orders': _pending_orders(),
    })


@_staff_required
def manage_add_blog(request):
    error = None
    if request.method == 'POST':
        title   = request.POST.get('title', '').strip()
        excerpt = request.POST.get('excerpt', '').strip()
        content = request.POST.get('content', '').strip()
        author  = request.POST.get('author', 'Malola Team').strip()
        published = request.POST.get('is_published') == 'on'
        cover   = request.FILES.get('cover_image')
        if not title:
            error = 'Title is required.'
        elif not content:
            error = 'Content is required.'
        else:
            post = BlogPost(title=title, excerpt=excerpt, content=content,
                            author=author, is_published=published)
            if cover:
                post.cover_image = cover
            post.save()
            return redirect('manage_blog')
    return render(request, 'store/admin_blog_form.html', {
        'action': 'Add', 'error': error,
        'pending_orders': _pending_orders(),
    })


@_staff_required
def manage_edit_blog(request, pk):
    post  = get_object_or_404(BlogPost, pk=pk)
    error = None
    if request.method == 'POST':
        post.title       = request.POST.get('title', '').strip()
        post.excerpt     = request.POST.get('excerpt', '').strip()
        post.content     = request.POST.get('content', '').strip()
        post.author      = request.POST.get('author', 'Malola Team').strip()
        post.is_published = request.POST.get('is_published') == 'on'
        cover = request.FILES.get('cover_image')
        if not post.title:
            error = 'Title is required.'
        elif not post.content:
            error = 'Content is required.'
        else:
            if cover:
                post.cover_image = cover
            post.save()
            return redirect('manage_blog')
    return render(request, 'store/admin_blog_form.html', {
        'action': 'Edit', 'post': post, 'error': error,
        'pending_orders': _pending_orders(),
    })


@_staff_required
@require_POST
def manage_delete_blog(request, pk):
    post = get_object_or_404(BlogPost, pk=pk)
    if post.cover_image:
        post.cover_image.delete(save=False)
    post.delete()
    return redirect('manage_blog')


@_staff_required
@require_POST
def manage_toggle_blog(request, pk):
    post = get_object_or_404(BlogPost, pk=pk)
    post.is_published = not post.is_published
    post.save(update_fields=['is_published'])
    return redirect('manage_blog')


def submit_review(request, order_pk):
    if not request.user.is_authenticated:
        return redirect(f'/?login=1')
    order = get_object_or_404(Order, pk=order_pk)
    if order.user and order.user != request.user:
        return HttpResponseForbidden(
            '<h2 style="font-family:sans-serif;padding:40px">403 — This review link does not belong to your account.</h2>'
        )
    def _ctx(**kw):
        c = _nav_ctx(); c['order'] = order; c.update(kw); return c
    already = hasattr(order, 'review') and order.review is not None
    if already:
        return render(request, 'store/submit_review.html', _ctx(already_reviewed=True))
    if request.method == 'POST':
        name   = request.POST.get('name', '').strip() or order.name
        text   = request.POST.get('review_text', '').strip()
        rating = request.POST.get('rating', '5')
        photo  = request.FILES.get('photo')
        if not text:
            return render(request, 'store/submit_review.html', _ctx(error='Please write your review before submitting.'))
        img_err = _validate_image(photo)
        if img_err:
            return render(request, 'store/submit_review.html', _ctx(error=img_err))
        if text:
            rev = Review(
                order=order, name=name, review_text=text,
                rating=min(5, max(1, int(rating) if rating.isdigit() else 5)),
                is_active=True,
            )
            if photo:
                rev.photo = photo
            rev.save()
            return render(request, 'store/submit_review.html', _ctx(success=True))
        return render(request, 'store/submit_review.html', _ctx(error='Please write your review before submitting.'))
    return render(request, 'store/submit_review.html', _ctx())


# ── checkout & payment ────────────────────────────────────────────────────────

def checkout_page(request, order_id):
    if not request.user.is_authenticated:
        return redirect(f'/?login=1')
    order = get_object_or_404(Order, id=order_id, user=request.user)
    if order.payment_status in ('paid', 'cod'):
        return redirect('orders')
    db_new_arrivals = Product.objects.filter(category='new_arrival', is_active=True)
    return render(request, 'store/checkout.html', {
        'order':            order,
        'razorpay_key':     settings.RAZORPAY_KEY_ID,
        'db_new_arrivals':  db_new_arrivals,
    })


@csrf_exempt
@require_POST
def create_razorpay_order(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Login required'}, status=401)
    try:
        data     = json.loads(request.body)
        order    = get_object_or_404(Order, id=data['order_id'], user=request.user)
        if order.payment_status in ('paid', 'cod'):
            return JsonResponse({'error': 'Order is already paid.'}, status=400)
        rp_order = _rp_create_order(int(float(order.total) * 100), f'order_{order.id}')
        order.razorpay_order_id = rp_order['id']
        order.save(update_fields=['razorpay_order_id'])
        return JsonResponse({
            'razorpay_order_id': rp_order['id'],
            'amount':            rp_order['amount'],
            'currency':          rp_order['currency'],
            'key':               settings.RAZORPAY_KEY_ID,
        })
    except urllib.error.URLError:
        return JsonResponse({'error': 'Could not connect to payment gateway. Check your Razorpay keys.'}, status=502)
    except Exception as exc:
        return JsonResponse({'error': str(exc)}, status=400)


@csrf_exempt
@require_POST
def verify_payment(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Login required'}, status=401)
    try:
        data = json.loads(request.body)
        ok   = _rp_verify(
            data['razorpay_order_id'],
            data['razorpay_payment_id'],
            data['razorpay_signature'],
        )
        if not ok:
            return JsonResponse({'success': False, 'error': 'Signature mismatch'}, status=400)
        order = get_object_or_404(Order, id=data['order_db_id'], user=request.user)
        if order.payment_status == 'paid':
            return JsonResponse({'success': True})   # webhook already confirmed it
        amount_paise = int(float(order.total) * 100)
        _confirm_order_paid(data['razorpay_order_id'], data['razorpay_payment_id'], amount_paise)
        return JsonResponse({'success': True})
    except Exception as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)


@csrf_exempt
@require_POST
def razorpay_webhook(request):
    """
    Razorpay calls this endpoint directly on every payment event.
    This is the safety net — it works even if the user closes the browser tab.
    Configure the URL in: Razorpay Dashboard → Settings → Webhooks
    """
    sig = request.META.get('HTTP_X_RAZORPAY_SIGNATURE', '')
    if not _rp_verify_webhook(request.body, sig):
        return JsonResponse({'error': 'Invalid signature'}, status=400)

    try:
        event = json.loads(request.body)
        ev    = event.get('event', '')

        if ev == 'payment.captured':
            payload       = event['payload']['payment']['entity']
            rp_order_id   = payload.get('order_id', '')
            rp_payment_id = payload.get('id', '')
            amount_paise  = payload.get('amount', 0)
            _confirm_order_paid(rp_order_id, rp_payment_id, amount_paise)

        elif ev == 'payment.failed':
            payload     = event['payload']['payment']['entity']
            rp_order_id = payload.get('order_id', '')
            order = Order.objects.filter(razorpay_order_id=rp_order_id).first()
            if order and order.payment_status == 'pending':
                order.payment_status = 'failed'
                order.save(update_fields=['payment_status'])

    except Exception:
        pass  # never return non-200 to Razorpay after signature check — it will retry forever

    return JsonResponse({'status': 'ok'})


@csrf_exempt
@require_POST
def record_payment_failure(request):
    """Called by the frontend when Razorpay's popup reports a failure."""
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False}, status=401)
    try:
        data  = json.loads(request.body)
        order = Order.objects.filter(id=data['order_id'], user=request.user, payment_status='pending').first()
        if order:
            order.payment_status = 'failed'
            order.save(update_fields=['payment_status'])
    except Exception:
        pass
    return JsonResponse({'ok': True})


@csrf_exempt
@require_POST
def confirm_cod(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Login required'}, status=401)
    try:
        data  = json.loads(request.body)
        order = get_object_or_404(Order, id=data['order_id'], user=request.user)
        prev = order.status
        order.payment_method = 'cod'
        order.payment_status = 'cod'
        order.status         = 'confirmed'
        order.save(update_fields=['payment_method','payment_status','status'])
        _log_status(order, 'confirmed', request.user, note='COD order confirmed', from_status=prev)
        return JsonResponse({'success': True})
    except Exception as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)


@csrf_exempt
@require_POST
def cancel_order(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Login required'}, status=401)
    try:
        data  = json.loads(request.body)
        order = get_object_or_404(Order, id=data['order_id'], user=request.user)
        CANCELLABLE = ('pending', 'confirmed', 'processing')
        if order.status not in CANCELLABLE:
            return JsonResponse({'success': False, 'error': f'Orders that are "{order.status}" cannot be cancelled.'}, status=400)

        refund_note = ''
        with transaction.atomic():
            # Lock the order row to avoid a double-cancel race
            order = Order.objects.select_for_update().get(pk=order.pk)
            if order.status not in CANCELLABLE:
                return JsonResponse({'success': False, 'error': 'Order is no longer cancellable.'}, status=400)

            # 1) Restore stock for inventory-tracked products (mirror the checkout decrement)
            for item in order.items:
                slug = item.get('slug', '')
                qty  = int(item.get('qty', 1) or 1)
                Product.objects.filter(slug=slug, track_inventory=True).update(
                    stock_quantity=F('stock_quantity') + qty
                )

            # 2) Release the coupon hold (once)
            if order.coupon_id and not order.coupon_released:
                Coupon.objects.filter(pk=order.coupon_id, used_count__gt=0).update(
                    used_count=F('used_count') - 1
                )
                order.coupon_released = True

            # 3) Refund if the order was actually paid online
            new_payment_status = order.payment_status
            if order.payment_status == 'paid' and order.razorpay_payment_id:
                amount_paise = int(round(float(order.total) * 100))
                if amount_paise > 0:
                    refund, err = _rp_refund(order.razorpay_payment_id, amount_paise)
                    if refund and refund.get('id'):
                        order.razorpay_refund_id = refund['id']
                        new_payment_status = 'refunded'
                        refund_note = 'Refund initiated successfully.'
                    else:
                        # Do NOT silently fail — cancel the order but flag for manual review
                        new_payment_status = 'refund_failed'
                        refund_note = f'Refund FAILED ({err}). Flagged for manual review.'
                        _alert_admin_refund_failure(order, err)
                else:
                    new_payment_status = 'refunded'
            # 'cod' / 'free' / 'pending' / 'failed' → nothing to refund

            order.payment_status = new_payment_status
            order.status         = 'cancelled'
            order.save(update_fields=['status', 'payment_status', 'razorpay_refund_id', 'coupon_released'])
            _log_status(order, 'cancelled', request.user, note=refund_note or 'Order cancelled by customer.', from_status='confirmed')

        return JsonResponse({
            'success': True,
            'refunded': order.payment_status == 'refunded',
            'refund_pending': order.payment_status == 'refund_failed',
            'message': refund_note,
        })
    except Exception as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)


def order_invoice(request, order_id):
    """Download a PDF invoice. Customer can fetch their own; staff can fetch any."""
    if not request.user.is_authenticated:
        return redirect('/?login=1')
    if request.user.is_staff:
        order = get_object_or_404(Order, id=order_id)
    else:
        order = get_object_or_404(Order, id=order_id, user=request.user)
    from .invoice import build_invoice_pdf
    from django.http import HttpResponse
    pdf = build_invoice_pdf(order)
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="Malola-Invoice-ORD{order.id:04d}.pdf"'
    return resp


def _alert_admin_refund_failure(order, err):
    """Email staff when an automatic refund fails so it can be handled manually."""
    try:
        from django.core.mail import mail_admins
        mail_admins(
            subject=f'[Malola] Refund FAILED for Order #{order.id}',
            message=(f'Automatic Razorpay refund failed for Order #{order.id} '
                     f'(₹{order.total}, payment {order.razorpay_payment_id}).\n'
                     f'Error: {err}\n\nPlease refund manually from the Razorpay dashboard.'),
            fail_silently=True,
        )
    except Exception:
        pass


# ── order API ─────────────────────────────────────────────────────────────────

@csrf_exempt
@require_POST
def place_order(request):
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Login required'}, status=401)
    # Rate limit: max 5 checkout attempts per user per 10 minutes
    if _is_rate_limited(f'checkout:{request.user.id}', 5, 600):
        return JsonResponse({'success': False, 'error': 'Too many checkout attempts. Please wait a few minutes and try again.', 'code': 'rate_limited'}, status=429)
    try:
        data  = json.loads(request.body)
        name    = data.get('name', '').strip()
        phone   = data.get('phone', '').strip()
        address = data.get('address', '').strip()
        if not name or not phone or not address:
            return JsonResponse({'success': False, 'error': 'Name, phone and address are required.'}, status=400)

        raw_items = data.get('items', [])
        if not raw_items:
            return JsonResponse({'success': False, 'error': 'Cart is empty.'}, status=400)

        coupon_code = data.get('coupon_code', '').strip().upper()

        # ── Atomic, row-locked stock check → order create → stock decrement ──
        # select_for_update() locks each Product row for the duration of the
        # transaction, so two concurrent orders for the last unit cannot both pass.
        with transaction.atomic():
            verified_items = []
            to_decrement   = []   # [(product, qty)] for inventory-tracked products
            total = 0
            gross_gst = 0.0       # GST component of the inclusive line totals (pre-discount)
            for item in raw_items:
                slug = item.get('slug', '').strip()
                try:
                    qty = int(item.get('qty', 1))
                except (TypeError, ValueError):
                    return JsonResponse({'success': False, 'error': 'invalid_input'}, status=400)
                if qty < 1:
                    return JsonResponse({'success': False, 'error': 'invalid_input'}, status=400)
                if not slug:
                    return JsonResponse({'success': False, 'error': f'Product "{item.get("name","")}" has no slug.'}, status=400)

                product = Product.objects.select_for_update().filter(slug=slug, is_active=True).first()
                if not product:
                    return JsonResponse({'success': False, 'error': f'Product "{item.get("name","")}" not found or unavailable.', 'code': 'product_unavailable'}, status=400)

                # Per-order quantity cap
                if product.max_order_qty and qty > product.max_order_qty:
                    return JsonResponse({'success': False, 'error': f'You can order at most {product.max_order_qty} of "{product.title}".', 'code': 'max_qty_reached', 'max': product.max_order_qty}, status=400)

                # Inventory enforcement (only when this product tracks stock)
                if product.track_inventory:
                    if product.stock_quantity <= 0:
                        return JsonResponse({'success': False, 'error': f'"{product.title}" is out of stock.', 'code': 'out_of_stock'}, status=400)
                    if qty > product.stock_quantity:
                        return JsonResponse({'success': False, 'error': f'"{product.title}" only has {product.stock_quantity} unit(s) left in stock.', 'code': 'insufficient_stock', 'available': product.stock_quantity}, status=400)
                    to_decrement.append((product, qty))

                # Price is server-authoritative. For a selected weight variant we
                # scale the BASE price by grams/base_grams (validated against the
                # product's own weight list) — never trust the client's price.
                base_price = float(product.price)
                price = base_price
                weight = (item.get('weight') or '').strip()
                if weight:
                    allowed = [w.strip() for w in (product.weights or '').split(',') if w.strip()]
                    if weight in allowed:
                        grams      = _parse_grams(weight)
                        base_grams = _parse_grams(allowed[0]) if allowed else grams
                        if grams and base_grams:
                            price = round(base_price * grams / base_grams, 2)
                    else:
                        weight = ''   # unknown variant → ignore, charge base price
                rate  = float(product.gst_rate or 0)
                line  = price * qty
                total += line
                # GST-inclusive: tax component = line * rate / (100 + rate)
                gross_gst += (line * rate / (100 + rate)) if rate else 0.0
                verified_items.append({
                    'name':     product.title + (f' ({weight})' if weight else ''),
                    'slug':     product.slug,
                    'price':    price,
                    'qty':      qty,
                    'weight':   weight,
                    'gst_rate': rate,   # snapshot — old orders keep their rate if admin changes it later
                    'image':    item.get('image', ''),
                })

            # ── Coupon validation (locked so used_count can't be over-redeemed) ──
            discount_amount = 0.0
            coupon_obj      = None
            if coupon_code:
                from django.utils import timezone
                coupon_obj = Coupon.objects.select_for_update().filter(
                    code=coupon_code, is_active=True,
                    valid_from__lte=timezone.now(), valid_until__gte=timezone.now(),
                ).first()
                if coupon_obj and coupon_obj.used_count < coupon_obj.max_uses:
                    discount_amount = coupon_obj.calculate_discount(total)
                else:
                    coupon_obj = None

            final_total = max(0, round(total - discount_amount, 2))

            # GST on the amount actually charged (scale the inclusive tax by the discount ratio)
            scale       = (final_total / total) if total > 0 else 0
            gst_amount  = round(gross_gst * scale, 2)
            subtotal    = round(final_total - gst_amount, 2)

            # ₹0 order (e.g. 100%-off coupon): skip Razorpay, confirm directly
            is_free = final_total <= 0

            order = Order.objects.create(
                user            = request.user,
                name            = name,
                phone           = phone,
                address         = address,
                items           = verified_items,
                total           = final_total,
                subtotal        = subtotal,
                gst_amount      = gst_amount,
                payment_status  = 'paid' if is_free else 'pending',
                payment_method  = 'free' if is_free else '',
                status          = 'confirmed' if is_free else 'pending',
                coupon          = coupon_obj,
                coupon_code     = coupon_code if coupon_obj else '',
                discount_amount = round(discount_amount, 2),
            )
            if coupon_obj:
                Coupon.objects.filter(pk=coupon_obj.pk).update(used_count=F('used_count') + 1)
            # Decrement stock on the rows we locked above
            for product, qty in to_decrement:
                Product.objects.filter(pk=product.pk).update(stock_quantity=F('stock_quantity') - qty)
            if is_free:
                _log_status(order, 'confirmed', request.user, note='Free order (₹0) auto-confirmed', from_status='pending')

        # Did the authoritative total differ from what the cart showed the customer?
        try:
            client_total = float(data.get('total', 0))
        except (TypeError, ValueError):
            client_total = 0
        price_changed = client_total > 0 and abs(client_total - float(final_total)) > 0.5

        if is_free:
            return JsonResponse({
                'success':  True,
                'order_id': order.id,
                'redirect': '/orders/',
                'free':     True,
            })
        return JsonResponse({
            'success':       True,
            'order_id':      order.id,
            'redirect':      f'/checkout/{order.id}/',
            'price_changed': price_changed,
            'server_total':  float(final_total),
        })
    except (KeyError, ValueError, json.JSONDecodeError) as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_GET
def product_list_api(request):
    """All active products for the search widget — slug, name, price, category, image."""
    qs = Product.objects.filter(is_active=True).values(
        'slug', 'title', 'price', 'discount_price', 'product_type', 'image'
    )
    results = []
    for p in qs:
        results.append({
            'id':    p['slug'],
            'name':  p['title'],
            'price': float(p['price']),
            'mrp':   float(p['discount_price']) if p['discount_price'] else None,
            'cat':   p['product_type'] or 'Snack',
            'img':   settings.MEDIA_URL + p['image'] if p['image'] else '',
        })
    return JsonResponse({'products': results})


@require_GET
def get_saved_address(request):
    """Return the name/phone/address from the user's most recent order."""
    if not request.user.is_authenticated:
        return JsonResponse({'found': False})
    order = Order.objects.filter(user=request.user).order_by('-created_at').first()
    if not order:
        return JsonResponse({'found': False})
    return JsonResponse({
        'found':   True,
        'name':    order.name,
        'phone':   order.phone,
        'address': order.address,
    })


@require_GET
def get_order(request, order_id):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Login required'}, status=401)
    order = get_object_or_404(Order, id=order_id, user=request.user)
    return JsonResponse({
        'id':         order.id,
        'name':       order.name,
        'phone':      order.phone,
        'address':    order.address,
        'items':      order.items,
        'total':      str(order.total),
        'status':     order.status,
        'created_at': order.created_at.isoformat(),
    })


# ── auth API ──────────────────────────────────────────────────────────────────

@require_POST
def register_view(request):
    from django.contrib.auth.password_validation import validate_password
    from django.core.exceptions import ValidationError
    from django.db import IntegrityError
    ip = _client_ip(request)
    if _is_rate_limited(f'register:{ip}', max_hits=5, window_seconds=3600):
        return JsonResponse({'success': False, 'error': 'Too many attempts. Try again in an hour.'}, status=429)
    try:
        data     = json.loads(request.body)
        name     = data.get('name', '').strip()
        email    = data.get('email', '').strip().lower()
        password = data.get('password', '')
        if not email or not password:
            return JsonResponse({'success': False, 'error': 'Email and password are required'})
        if '@' not in email or '.' not in email.split('@')[-1]:
            return JsonResponse({'success': False, 'error': 'Please enter a valid email address'})
        if User.objects.filter(username=email).exists():
            return JsonResponse({'success': False, 'error': 'Email is already registered'})
        parts      = name.split()
        first_name = parts[0] if parts else ''
        last_name  = ' '.join(parts[1:]) if len(parts) > 1 else ''
        # Enforce Django's AUTH_PASSWORD_VALIDATORS (length, common, numeric, similarity)
        try:
            validate_password(password, User(username=email, email=email, first_name=first_name))
        except ValidationError as e:
            return JsonResponse({'success': False, 'error': ' '.join(e.messages)})
        try:
            user = User.objects.create_user(
                username=email, email=email, password=password,
                first_name=first_name, last_name=last_name,
            )
        except IntegrityError:
            # Race: another request registered the same email between the check and create
            return JsonResponse({'success': False, 'error': 'Email is already registered'})
        auth_login(request, user)
        return JsonResponse({'success': True, 'name': name or email})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid data'}, status=400)


@require_POST
def login_view(request):
    ip = _client_ip(request)
    if _is_rate_limited(f'login:{ip}', max_hits=10, window_seconds=300):
        return JsonResponse({'success': False, 'error': 'Too many login attempts. Try again in 5 minutes.'}, status=429)
    try:
        data     = json.loads(request.body)
        email    = data.get('email', '').strip().lower()
        password = data.get('password', '')
        user     = authenticate(request, username=email, password=password)
        if user:
            auth_login(request, user)
            name = user.get_full_name() or user.first_name or user.username
            return JsonResponse({'success': True, 'name': name})
        return JsonResponse({'success': False, 'error': 'Invalid email or password'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid data'}, status=400)


@require_POST
def logout_view(request):
    auth_logout(request)
    return JsonResponse({'success': True})


# ── Coupon API ────────────────────────────────────────────────────────────────

@csrf_exempt
@require_POST
def apply_coupon(request):
    """Validate a coupon code for a given cart total and return the discount."""
    if not request.user.is_authenticated:
        return JsonResponse({'valid': False, 'error': 'Login required'}, status=401)
    try:
        from django.utils import timezone
        d           = json.loads(request.body)
        code        = d.get('code', '').strip().upper()
        cart_total  = float(d.get('cart_total', 0))
        if not code:
            return JsonResponse({'valid': False, 'error': 'Coupon code is required'}, status=400)
        coupon = Coupon.objects.filter(
            code=code, is_active=True,
            valid_from__lte=timezone.now(),
            valid_until__gte=timezone.now(),
        ).first()
        if not coupon:
            return JsonResponse({'valid': False, 'error': 'Invalid or expired coupon'})
        if coupon.used_count >= coupon.max_uses:
            return JsonResponse({'valid': False, 'error': 'Coupon usage limit reached'})
        if cart_total < float(coupon.min_order_value):
            return JsonResponse({
                'valid': False,
                'error': f'Minimum order value ₹{coupon.min_order_value} required for this coupon',
            })
        discount = coupon.calculate_discount(cart_total)
        return JsonResponse({
            'valid':            True,
            'discount_amount':  discount,
            'new_total':        round(cart_total - discount, 2),
            'description':      coupon.description or str(coupon),
        })
    except Exception as exc:
        return JsonResponse({'valid': False, 'error': str(exc)}, status=400)


# ── DB Cart sync ──────────────────────────────────────────────────────────────

@csrf_exempt
@require_POST
def sync_cart(request):
    """Sync the localStorage cart to the server-side Cart model."""
    if not request.user.is_authenticated:
        return JsonResponse({'ok': True})   # silently skip for guests
    try:
        items = json.loads(request.body).get('items', [])
        cart, _ = Cart.objects.get_or_create(user=request.user)
        cart.items.all().delete()
        for item in items:
            slug = item.get('slug', '').strip()
            qty  = max(1, int(item.get('qty', 1)))
            if not slug:
                continue
            product = Product.objects.filter(slug=slug, is_active=True).first()
            if product:
                CartItem.objects.create(cart=cart, product=product, qty=qty)
        return JsonResponse({'ok': True})
    except Exception as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=400)


@require_GET
def get_db_cart(request):
    """Return the server-side cart for the logged-in user."""
    if not request.user.is_authenticated:
        return JsonResponse({'items': [], 'total': 0})
    cart = Cart.objects.filter(user=request.user).prefetch_related('items__product').first()
    if not cart:
        return JsonResponse({'items': [], 'total': 0})
    items = [
        {
            'name':  ci.product.title,
            'slug':  ci.product.slug,
            'price': float(ci.product.discount_price or ci.product.price),
            'qty':   ci.qty,
            'image': ci.product.image.url if ci.product.image else '',
        }
        for ci in cart.items.all()
    ]
    return JsonResponse({'items': items, 'total': cart.get_total()})


# ── Password reset (OTP-based) ────────────────────────────────────────────────

@csrf_exempt
@require_POST
def request_password_reset(request):
    ip = _client_ip(request)
    if _is_rate_limited(f'pw_reset:{ip}', max_hits=5, window_seconds=3600):
        return JsonResponse({'ok': False, 'error': 'Too many attempts. Try again in an hour.'}, status=429)
    try:
        import random
        d     = json.loads(request.body)
        email = d.get('email', '').strip().lower()
        if not email:
            return JsonResponse({'ok': False, 'error': 'Email is required'}, status=400)
        user = User.objects.filter(email=email).first() or User.objects.filter(username=email).first()
        if not user:
            # Return ok to avoid user enumeration
            return JsonResponse({'ok': True, 'message': 'If that email exists, an OTP has been sent.'})
        otp = str(random.randint(100000, 999999))
        cache.set(f'pw_reset_otp:{email}', otp, timeout=600)  # 10 min
        try:
            from notifications.email import send_password_reset_otp
            send_password_reset_otp(user, otp)
        except Exception:
            pass
        return JsonResponse({'ok': True, 'message': 'If that email exists, an OTP has been sent.'})
    except Exception as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=400)


@csrf_exempt
@require_POST
def confirm_password_reset(request):
    from django.contrib.auth.password_validation import validate_password
    from django.core.exceptions import ValidationError
    ip = _client_ip(request)
    # Backstop IP rate limit on the confirm endpoint itself
    if _is_rate_limited(f'pw_confirm:{ip}', max_hits=10, window_seconds=600):
        return JsonResponse({'ok': False, 'error': 'Too many attempts. Try again later.'}, status=429)
    try:
        d        = json.loads(request.body)
        email    = d.get('email', '').strip().lower()
        otp      = d.get('otp', '').strip()
        new_pass = d.get('new_password', '')
        if not email or not otp or not new_pass:
            return JsonResponse({'ok': False, 'error': 'email, otp and new_password are required'}, status=400)

        stored_otp = cache.get(f'pw_reset_otp:{email}')
        if not stored_otp:
            return JsonResponse({'ok': False, 'error': 'Invalid or expired OTP. Please request a new one.'}, status=400)

        # Cap OTP guesses: 5 wrong tries → invalidate the OTP (defeats brute force)
        attempts_key = f'pw_reset_attempts:{email}'
        if stored_otp != otp:
            attempts = (cache.get(attempts_key, 0)) + 1
            cache.set(attempts_key, attempts, timeout=600)
            if attempts >= 5:
                cache.delete(f'pw_reset_otp:{email}')
                cache.delete(attempts_key)
                return JsonResponse({'ok': False, 'error': 'Too many incorrect attempts. Please request a new OTP.'}, status=400)
            return JsonResponse({'ok': False, 'error': 'Invalid or expired OTP'}, status=400)

        user = User.objects.filter(email=email).first() or User.objects.filter(username=email).first()
        if not user:
            return JsonResponse({'ok': False, 'error': 'User not found'}, status=404)
        # Enforce Django's password validators
        try:
            validate_password(new_pass, user)
        except ValidationError as e:
            return JsonResponse({'ok': False, 'error': ' '.join(e.messages)}, status=400)
        user.set_password(new_pass)
        user.save()
        cache.delete(f'pw_reset_otp:{email}')
        cache.delete(attempts_key)
        return JsonResponse({'ok': True, 'message': 'Password reset successfully. Please log in.'})
    except Exception as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=400)


# ── Admin — Coupons ───────────────────────────────────────────────────────────

@_staff_required
def manage_coupons(request):
    coupons = Coupon.objects.all()
    return render(request, 'store/admin_coupons.html', {
        'coupons':       coupons,
        'pending_orders': _pending_orders(),
    })


@_staff_required
def manage_add_coupon(request):
    from django.utils import timezone
    error = None
    if request.method == 'POST':
        code            = request.POST.get('code', '').strip().upper()
        description     = request.POST.get('description', '').strip()
        discount_type   = request.POST.get('discount_type', 'percent')
        discount_value  = request.POST.get('discount_value', '0')
        min_order_value = request.POST.get('min_order_value', '0')
        max_discount    = request.POST.get('max_discount', '').strip()
        valid_from      = request.POST.get('valid_from', '')
        valid_until     = request.POST.get('valid_until', '')
        max_uses        = request.POST.get('max_uses', '100')
        is_active       = request.POST.get('is_active') == 'on'
        if not code:
            error = 'Coupon code is required.'
        elif Coupon.objects.filter(code=code).exists():
            error = f'Coupon code "{code}" already exists.'
        else:
            try:
                Coupon.objects.create(
                    code            = code,
                    description     = description,
                    discount_type   = discount_type,
                    discount_value  = float(discount_value),
                    min_order_value = float(min_order_value),
                    max_discount    = float(max_discount) if max_discount else None,
                    valid_from      = valid_from,
                    valid_until     = valid_until,
                    max_uses        = int(max_uses),
                    is_active       = is_active,
                )
                return redirect('manage_coupons')
            except Exception as exc:
                error = str(exc)
    return render(request, 'store/admin_coupon_form.html', {
        'action': 'Add', 'error': error, 'coupon': None,
        'pending_orders': _pending_orders(),
    })


@_staff_required
def manage_edit_coupon(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    error  = None
    if request.method == 'POST':
        coupon.code            = request.POST.get('code', '').strip().upper()
        coupon.description     = request.POST.get('description', '').strip()
        coupon.discount_type   = request.POST.get('discount_type', 'percent')
        coupon.discount_value  = float(request.POST.get('discount_value', 0))
        coupon.min_order_value = float(request.POST.get('min_order_value', 0))
        max_d = request.POST.get('max_discount', '').strip()
        coupon.max_discount    = float(max_d) if max_d else None
        coupon.valid_from      = request.POST.get('valid_from', '')
        coupon.valid_until     = request.POST.get('valid_until', '')
        coupon.max_uses        = int(request.POST.get('max_uses', 100))
        coupon.is_active       = request.POST.get('is_active') == 'on'
        try:
            coupon.save()
            return redirect('manage_coupons')
        except Exception as exc:
            error = str(exc)
    return render(request, 'store/admin_coupon_form.html', {
        'action': 'Edit', 'error': error, 'coupon': coupon,
        'pending_orders': _pending_orders(),
    })


@_staff_required
def manage_delete_coupon(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    if request.method == 'POST':
        coupon.delete()
    return redirect('manage_coupons')


# ── Wishlist ──────────────────────────────────────────────────────────────────

@require_GET
def get_wishlist(request):
    """Return slugs the user has wishlisted (stored in session)."""
    wishlist = request.session.get('wishlist', [])
    return JsonResponse({'slugs': wishlist})


@csrf_exempt
@require_POST
def toggle_wishlist(request):
    """Add/remove a product slug from the session-based wishlist."""
    try:
        slug = json.loads(request.body).get('slug', '').strip()
        if not slug:
            return JsonResponse({'error': 'slug required'}, status=400)
        if not Product.objects.filter(slug=slug, is_active=True).exists():
            return JsonResponse({'error': 'Product not found'}, status=404)
        wishlist = request.session.get('wishlist', [])
        if slug in wishlist:
            wishlist.remove(slug)
            added = False
        else:
            wishlist.append(slug)
            added = True
        request.session['wishlist'] = wishlist
        return JsonResponse({'added': added, 'total': len(wishlist)})
    except Exception as exc:
        return JsonResponse({'error': str(exc)}, status=400)


# ── Pincode serviceability ────────────────────────────────────────────────────

@require_GET
def check_pincode(request):
    """Check if a pincode is serviceable via Shiprocket."""
    pincode = request.GET.get('pincode', '').strip()
    if not pincode or len(pincode) != 6 or not pincode.isdigit():
        return JsonResponse({'serviceable': False, 'error': 'Invalid pincode'}, status=400)
    try:
        from shipping.shiprocket import get_serviceable_couriers
        data, err = get_serviceable_couriers(pincode)
        if err:
            return JsonResponse({'serviceable': True, 'note': 'Could not verify, proceed to order'})
        couriers = (data or {}).get('data', {}).get('available_courier_companies', [])
        return JsonResponse({'serviceable': len(couriers) > 0, 'couriers_count': len(couriers)})
    except Exception:
        return JsonResponse({'serviceable': True, 'note': 'Proceed to order'})


# ── Admin — Users ─────────────────────────────────────────────────────────────

@_staff_required
def manage_users(request):
    qs     = AuthUser.objects.select_related('profile').prefetch_related('orders').order_by('-date_joined')
    search = request.GET.get('q', '').strip()
    if search:
        qs = qs.filter(
            Q(username__icontains=search) | Q(email__icontains=search) |
            Q(first_name__icontains=search) | Q(last_name__icontains=search)
        )
    paginator = Paginator(qs, 25)
    users     = paginator.get_page(request.GET.get('page'))
    return render(request, 'store/admin_users.html', {
        'users': users, 'search': search,
        'pending_orders': _pending_orders(),
        'total_users': AuthUser.objects.count(),
    })


@_staff_required
def manage_user_detail(request, pk):
    from accounts.models import UserProfile, Address
    profile_user = get_object_or_404(AuthUser, pk=pk)
    orders    = Order.objects.filter(user=profile_user).order_by('-created_at')
    addresses = Address.objects.filter(user=profile_user)
    cart      = Cart.objects.filter(user=profile_user).prefetch_related('items__product').first()
    profile, _ = UserProfile.objects.get_or_create(user=profile_user)
    return render(request, 'store/admin_user_detail.html', {
        'profile_user': profile_user, 'profile': profile,
        'orders': orders, 'addresses': addresses, 'cart': cart,
        'pending_orders': _pending_orders(),
    })


@csrf_exempt
@require_POST
def manage_user_toggle_staff(request, pk):
    if not request.user.is_staff:
        return JsonResponse({'error': 'Admin only'}, status=403)
    u = get_object_or_404(AuthUser, pk=pk)
    if u == request.user:
        return JsonResponse({'error': 'Cannot change your own staff status'}, status=400)
    u.is_staff = not u.is_staff
    u.save(update_fields=['is_staff'])
    return JsonResponse({'is_staff': u.is_staff})


@csrf_exempt
@require_POST
def manage_user_reset_password(request, pk):
    if not request.user.is_staff:
        return JsonResponse({'error': 'Admin only'}, status=403)
    try:
        d        = json.loads(request.body)
        new_pass = d.get('new_password', '').strip()
        if len(new_pass) < 8:
            return JsonResponse({'error': 'Password must be at least 8 characters'}, status=400)
        u = get_object_or_404(AuthUser, pk=pk)
        u.set_password(new_pass)
        u.save()
        return JsonResponse({'ok': True})
    except Exception as exc:
        return JsonResponse({'error': str(exc)}, status=400)


# ── Admin — Carts ─────────────────────────────────────────────────────────────

@_staff_required
def manage_carts(request):
    qs     = Cart.objects.select_related('user').prefetch_related('items__product').order_by('-updated_at')
    search = request.GET.get('q', '').strip()
    if search:
        qs = qs.filter(
            Q(user__username__icontains=search) | Q(user__email__icontains=search)
        )
    paginator = Paginator(qs, 25)
    carts     = paginator.get_page(request.GET.get('page'))
    return render(request, 'store/admin_carts.html', {
        'carts': carts, 'search': search,
        'pending_orders': _pending_orders(),
        'total_carts': Cart.objects.count(),
    })


# ── Admin — Shipments ─────────────────────────────────────────────────────────

@_staff_required
def manage_shipments(request):
    from shipping.models import Shipment
    qs = Shipment.objects.select_related('order__user').order_by('-created_at')
    paginator = Paginator(qs, 25)
    shipments = paginator.get_page(request.GET.get('page'))
    # Orders confirmed/paid but with no shipment record yet
    shipped_order_ids = Shipment.objects.values_list('order_id', flat=True)
    needs_ship = Order.objects.filter(
        status__in=['confirmed', 'processing'],
        payment_status='paid',
    ).exclude(id__in=shipped_order_ids).order_by('-created_at')
    return render(request, 'store/admin_shipments.html', {
        'shipments': shipments,
        'pending_orders': needs_ship,
        'pending_count': needs_ship.count(),
        'total_shipments': Shipment.objects.count(),
    })
